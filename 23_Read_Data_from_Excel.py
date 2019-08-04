import openpyxl

path = "C:/Users/Usuario/PycharmProjects/Selenium_with_Python/Selenium_excel_ex.xlsx"

workbook = openpyxl.load_workbook(path)

sheet = workbook.active  # get_sheet_by_name("Hoja1")

rows = sheet.max_row  # 13
cols = sheet.max_column  # 4

print(rows)
print(cols)

for r in range(1, rows + 1):
    for c in range(1, cols + 1):
        print(sheet.cell(row=r, column=c).value, end="        ")

    print()
